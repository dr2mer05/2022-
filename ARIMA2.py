from openpyxl import load_workbook
from pandas_datareader import data as pdr
import yfinance as yf
yf.pdr_override()

import pandas as pd
from statsmodels.tsa.arima.model import ARIMA
import matplotlib.pyplot as plt

AR = 1
DIFF = 1
MA = 0

save_dir_path = "C:/Users/dr2mer05/Desktop/2022창개연/ARIMA_AR/"  # 그래프를 저장할 디렉토리의 위치

DATA_STUDY = 253  # 데이터의 총 개수
data_in_month = [21, 20, 21, 21, 21, 21, 22, 21, 21, 23, 19, 22]  # 2020 1월~12월의 월별 데이터 개수
data_in_month.reverse()

DATA_TEST = 19  # 테스트 데이터의 개수

min_list = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  # 학습 데이터의 가간별 오차 최솟값의 개수

tickers = []  # Nasdaq 100 기업(99개 기업)의 Ticker

# Nasdaq 100 기업의 ticker 불러오기
wb = load_workbook("C:/Users/dr2mer05/Desktop/2022창개연/Nasdaq100.xlsx", data_only=True)
ws = wb["Sheet1"]
get_cells = ws['A2':'A100']
for row in get_cells:
  for cell in row:
      tickers.append(cell.value)


for ticker in tickers:  # 각 ticker(기업)에 대해 반복한다
    error_by_month_list = []  # 학습 데이터의 기간에 따른 오차

    study_raw_df = pdr.get_data_yahoo(ticker, start="2020-01-01", end="2020-12-31")  # 200101~201231의 주가 데이터 불러오기
    study_raw_list = study_raw_df['Close'].tolist()  # 종가를 리스트로 저장하기

    test_df = pdr.get_data_yahoo(ticker, start="2021-01-02", end="2021-02-01")  # 210101~210131의 주가 데이터 불러오기

    # 날짜(일)을 리스트로 저장하기(리스트의 인덱스는 0부터 시작하므로 날짜(일)-1)
    test_df['Date_datetime'] = pd.to_datetime(test_df.index.get_level_values('Date'))
    test_df['Date_index'] = test_df['Date_datetime'].dt.day-1
    test_date_list = test_df['Date_index'].values.tolist()

    test_list = test_df['Close'].values.tolist()  # 종가를 리스트로 저장하기

    data_study_num = 0  # 학습 데이터의 개수
    for dim in data_in_month:
        data_study_num += dim
        study_list = study_raw_list[DATA_STUDY-data_study_num:DATA_STUDY]
        # study_list = study_raw_list[DATA_STUDY-data_study_num:DATA_STUDY-data_study_num+dim]

        # ARIMA model 생성
        model = ARIMA(study_list, order=(AR, DIFF, MA))
        fit = model.fit()

        predict_list = fit.forecast(steps=31)  # 2021년 1월에 대해 예측하기

        # 상대 오차의 평균 계산
        error_list = []
        i = 0
        for test_date in test_date_list:
            error_list.append((predict_list[test_date]-test_list[i]) / predict_list[test_date] * 100)
            i += 1
        for i in range(DATA_TEST):
            error_list[i] = abs(error_list[i])
        error = sum(error_list) / len(error_list)
        error_by_month_list.append(error)

    # 학습 데이터의 기간에 따른 오차 그래프 그리기
    plt.plot(error_by_month_list)
    plt.savefig(save_dir_path+"error_"+ticker+".png")
    plt.close()

    temp = min(error_by_month_list)
    index = error_by_month_list.index(temp)
    min_list[index] += 1

# 학습 데이터의 가간별 오차 최솟값의 개수 그래프 그리기
plt.plot(min_list)
plt.savefig(save_dir_path+"min"+".png")



